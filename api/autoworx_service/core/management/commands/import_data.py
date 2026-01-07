from django.core.management.base import BaseCommand
from core.models import PredefinedCar, ServiceType, PredefinedTask, PredefinedPart, Part
from django.core.files import File
import openpyxl
import logging
import os

logger = logging.getLogger(__name__)


class Command(BaseCommand):
    help = 'Import data from Excel file'

    def handle(self, *args, **options):
        # Check if data already exists
        if ServiceType.objects.exists():
            self.stdout.write(self.style.SUCCESS('Data already imported. Skipping import.'))
            return
        
        excel_file = './core/management/commands/predefined_tasks.xlsx'  
        total_rows = self.count_total_rows(excel_file)
        processed_rows = 0
        
        try:
            workbook = openpyxl.load_workbook(excel_file)

            # ServiceTypes
            service_types_sheet = workbook['ServiceTypes']
            service_types_data = []
            for row in service_types_sheet.iter_rows(min_row=2, values_only=True):
                service_type_id, name, description, labour_cost = row
                service_types_data.append(ServiceType(name=name, description=description, labour_cost=labour_cost))
            ServiceType.objects.bulk_create(service_types_data)

            # PredefinedParts
            predefined_parts_sheet = workbook['PredefinedParts']
            predefined_parts_data = []
            for row in predefined_parts_sheet.iter_rows(min_row=2, values_only=True):
                part_id, name, description, service_type_id = row
                service_type = ServiceType.objects.get(id=service_type_id)
                predefined_parts_data.append(PredefinedPart(name=name, description=description, service_type=service_type))
            PredefinedPart.objects.bulk_create(predefined_parts_data)

            # PredefinedTasks
            predefined_tasks_sheet = workbook['PredefinedTasks']
            predefined_tasks_data = []
            for row in predefined_tasks_sheet.iter_rows(min_row=2, values_only=True):
                task_id, service_type_id, description = row
                service_type = ServiceType.objects.get(id=service_type_id)
                predefined_tasks_data.append(PredefinedTask(service_type=service_type, description=description))
            PredefinedTask.objects.bulk_create(predefined_tasks_data)

            # PredefinedCars
            predefined_cars_sheet = workbook['PredefinedCars']
            for row in predefined_cars_sheet.iter_rows(min_row=2, values_only=True):
                make, model, year, image_path = row
                image_full_path = os.path.join('media', image_path)
                
                if os.path.exists(image_full_path):
                    with open(image_full_path, 'rb') as img_file:
                        car = PredefinedCar(make=make, model=model, year=year)
                        car.image.save(os.path.basename(image_path), File(img_file), save=True)
                else:
                    logger.warning(f"Image file not found: {image_full_path}")
                    car = PredefinedCar.objects.create(make=make, model=model, year=year)
                
                processed_rows += 1
                self.report_progress(processed_rows, total_rows)

            # Parts
            parts_sheet = workbook['Parts']
            for row in parts_sheet.iter_rows(min_row=2, values_only=True):
                part_id, predefined_car_make, predefined_car_model, predefined_car_year, predefined_part_id, price = row
                predefined_car = PredefinedCar.objects.get(make=predefined_car_make, model=predefined_car_model, year=predefined_car_year)
                predefined_part = PredefinedPart.objects.get(id=predefined_part_id)
                Part.objects.create(predefined_car=predefined_car, predefined_part=predefined_part, price=price)
                
            self.stdout.write(self.style.SUCCESS('Import completed successfully!'))
        except Exception as e:
            logger.error(f"Error importing data: {e}")
            self.stdout.write(self.style.ERROR(f"An error occurred during import: {e}"))

    def count_total_rows(self, excel_file):
        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            return sheet.max_row - 1  # Exclude header row
        except Exception as e:
            logger.error(f"Error counting rows in excel file: {e}")
            return 0  # Handle error gracefully

    def report_progress(self, processed, total):
        if total > 0:
            percentage = int((processed / total) * 100)
            self.stdout.write(self.style.WARNING(f"Progress: {processed}/{total} ({percentage}%)"))
        